#!/usr/bin/env python3
"""
Phase 0A: Legacy XLSX Migration to Supabase contractor_rows

Reads all source sheets from the legacy REFORM AI workbooks, performs a
canonical dedupe audit, merges with priority order:
    Contractors enriched > Contractors master > Request Failed
then upserts every record into Supabase contractor_rows.

Catalog Classification rows enrich matching (company, subcategory) records.
Each legacy company-subcategory pair is preserved as a separate row.

Run BEFORE enabling any automated orchestrator runs.
Run --dry-run first to validate counts without writing to Supabase.

Usage:
    python migration/migrate_legacy_xlsx.py --dry-run
    python migration/migrate_legacy_xlsx.py
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

# ---------------------------------------------------------------------------
# Resolve project root and load .env.local
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import find_dotenv, load_dotenv

load_dotenv(find_dotenv(".env.local", usecwd=True))

from supabase import create_client

# ---------------------------------------------------------------------------
# Source paths — read from new location, keep originals untouched
# ---------------------------------------------------------------------------
PIPELINE_WORKBOOK = Path(
    r"C:\Users\cjlea\AI-Projects\ReformAI_Agents\Shared Docs\REFORM AI B2B PIPELINE (official).xlsx"
)
CATALOG_WORKBOOK = Path(
    r"C:\Users\cjlea\AI-Projects\ReformAI_Agents\Shared Docs\Master_ReformAI_Contractor_Catalogs.xlsx"
)

REPORT_DIR = SCRIPT_DIR / "reports"

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MARKET_ID = "co-renovation"
MIGRATION_BATCH_ID = f"migration_{date.today().isoformat()}"

MASTER_SHEET   = "Contractors master"
ENRICHED_SHEET = "Contractors enriched"
FAILED_SHEET   = "Request Failed"
LEGACY_SHEET   = "Contractors"
CATALOG_SHEET  = "Catalog Classification"

# Derived / review-only sheets — never used as migration source
SKIP_SHEETS = {
    "Pipeline Summary",
    "Provider Tab Summary",
    "Catalog Classification",  # handled separately from catalog workbook
}

# Legacy extraction status values → canonical
STATUS_MAP = {
    "completed":    "extraction_completed",
    "not_found":    "no_contact_found",
    "missing_url":  "request_failed",
    "skipped":      "",
}

# Batch size for Supabase upserts
UPSERT_BATCH = 50


# ---------------------------------------------------------------------------
# Normalisation helpers
# ---------------------------------------------------------------------------

def norm_company(value: object) -> str:
    """Lowercase, strip accents, collapse whitespace."""
    text = str(value or "").strip()
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_only = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", ascii_only.lower().strip())


def norm_city(value: object) -> str:
    city_map = {
        "bogota":        "Bogota",
        "bogotá":        "Bogota",
        "medellin":      "Medellin",
        "medellín":      "Medellin",
        "cali":          "Cali",
        "barranquilla":  "Barranquilla",
        "santa marta":   "Santa Marta",
        "nacional":      "Nacional",
    }
    text = str(value or "").strip()
    nfkd = unicodedata.normalize("NFKD", text)
    folded = "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()
    return city_map.get(folded, text.strip())


def norm_domain(url: object) -> str:
    raw = str(url or "").strip().lower()
    if not raw:
        return ""
    parsed = urlparse(raw if "://" in raw else f"https://{raw}")
    domain = parsed.netloc or parsed.path
    return domain.removeprefix("www.").split("/")[0].strip()


def norm_status(value: object) -> str:
    raw = str(value or "").strip().lower()
    return STATUS_MAP.get(raw, raw)


def to_bool(value: object) -> bool | None:
    text = str(value or "").strip().lower()
    if text in ("yes", "true", "1"):
        return True
    if text in ("no", "false", "0"):
        return False
    return None  # Manual Review or blank


def to_int(value: object, default: int = 0) -> int:
    try:
        return int(float(str(value or default)))
    except (ValueError, TypeError):
        return default


def dedupe_key(company: str, subcategory: str, market_id: str) -> str:
    company_n = norm_company(company)
    sub_n = (subcategory or "").strip().lower()
    raw = f"{company_n}|{sub_n}|{market_id}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Sheet I/O
# ---------------------------------------------------------------------------

def list_sheets(path: Path) -> list[str]:
    if not path.exists():
        return []
    from openpyxl import load_workbook as _lw
    wb = _lw(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Return a list of row-dicts. Empty / whitespace-only rows are skipped.
    Each dict also carries __row_number__ (Excel data row, 2-based)."""
    if not path.exists():
        return []
    from openpyxl import load_workbook as _lw
    wb = _lw(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        wb.close()
        return []
    headers = [str(h or "").strip() for h in raw_headers]
    result = []
    for row_num, row in enumerate(rows_iter, start=2):
        d: dict[str, Any] = {"__row_number__": row_num}
        for i, h in enumerate(headers):
            d[h] = row[i] if i < len(row) else None
        if any(
            v is not None and str(v).strip()
            for k, v in d.items()
            if k != "__row_number__"
        ):
            result.append(d)
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Row mapping  (XLSX dict → contractor_rows insert dict)
# ---------------------------------------------------------------------------

def map_row(
    row: dict[str, Any],
    *,
    source_sheet: str,
    workbook_name: str,
) -> dict[str, Any]:
    company    = str(row.get("Company") or "").strip()
    subcategory = str(row.get("Subcategory") or "").strip()
    city_raw   = str(row.get("City") or "").strip()
    city       = norm_city(city_raw)
    url        = str(row.get("Company URL") or "").strip()
    status     = norm_status(row.get("Extraction Status"))

    return {
        # Market / identity
        "market_id":                   MARKET_ID,
        "company":                     norm_company(company),
        "company_display":             company or None,
        "city":                        city,
        "city_normalized":             city.lower() if city else "",
        "country":                     "CO",
        "category":                    str(row.get("Category") or "").strip() or None,
        "subcategory":                 subcategory or None,
        "integration_type":            str(row.get("Integration Type") or "").strip() or None,
        "role_in_reformai":            str(row.get("Role in Reform AI") or "").strip() or None,
        "company_url":                 url or None,
        "company_url_domain":          norm_domain(url) or None,
        # Dedupe keys
        "dedupe_key":                  dedupe_key(company, subcategory, MARKET_ID),
        "domain_dedupe_key":           norm_domain(url) or None,
        # Research
        "contact_name":                str(row.get("Contact Name") or "").strip() or None,
        "contact_role":                str(row.get("Role") or "").strip() or None,
        "contact_link":                str(row.get("Contact link") or "").strip() or None,
        "priority_tier":               str(row.get("Priority Tier") or "").strip() or None,
        "relationship_strength":       str(row.get("Relationship Strengthe") or "").strip() or None,
        "research_notes":              str(row.get("Notes") or "").strip() or None,
        "confidence_score":            None,
        # Extraction
        "email":                       str(row.get("Email") or "").strip() or None,
        "phone":                       str(row.get("Phone Number") or "").strip() or None,
        "contact_logic":               str(row.get("Contact Logic") or "").strip() or None,
        "contact_source_url":          str(row.get("Contact Source URL") or "").strip() or None,
        "extraction_status":           status or "pending",
        "extraction_notes":            str(row.get("Extraction Notes") or "").strip() or None,
        "failure_triage":              str(row.get("Failure Triage") or "").strip() or None,
        "validated_replacement_url":   str(row.get("Validated Replacement URL") or "").strip() or None,
        "validated_final_url":         str(row.get("Validated Final URL") or "").strip() or None,
        "url_validation_result":       str(row.get("URL Validation Result") or "").strip() or None,
        "search_discovery_result":     str(row.get("Search Discovery Result") or "").strip() or None,
        "search_replacement_url":      str(row.get("Search Replacement URL") or "").strip() or None,
        "search_discovery_notes":      str(row.get("Search Discovery Notes") or "").strip() or None,
        "search_source":               str(row.get("Search Source") or "").strip() or None,
        "search_query":                str(row.get("Search Query") or "").strip() or None,
        # Catalog (may be overridden by catalog enrichment step)
        "catalog_tier":                str(row.get("Catalog Tier") or "").strip() or None,
        "site_type":                   str(row.get("Site Type") or "").strip() or None,
        "catalog_confirmed":           to_bool(row.get("Catalog Confirmed")),
        "product_subpages_found":      to_bool(row.get("Product Subpages Found")),
        "product_count_estimate":      to_int(row.get("Product Count Estimate")),
        "image_count":                 to_int(row.get("Image Count")),
        "has_specs":                   to_bool(row.get("Has Specs")),
        "selected_for_extraction":     to_bool(row.get("Selected for Extraction")),
        "catalog_notes":               str(row.get("Catalog Notes") or "").strip() or None,
        # Migration metadata
        "source_system":               "legacy_xlsx",
        "migration_batch_id":          MIGRATION_BATCH_ID,
        "legacy_workbook_name":        workbook_name,
        "legacy_sheet_name":           source_sheet,
        "legacy_row_number":           row["__row_number__"],
        # HubSpot reserved
        "hubspot_sync_status":         "not_synced",
    }


# Extraction fields that can be overridden by enriched / failed sheets
_EXTRACTION_FIELDS = [
    "email", "phone", "contact_logic", "contact_source_url",
    "extraction_status", "extraction_notes", "failure_triage",
    "validated_replacement_url", "validated_final_url", "url_validation_result",
    "search_discovery_result", "search_replacement_url", "search_discovery_notes",
    "search_source", "search_query",
]

# Catalog fields that can be overridden by Catalog Classification sheet
_CATALOG_FIELDS = [
    "catalog_tier", "site_type", "catalog_confirmed",
    "product_subpages_found", "product_count_estimate",
    "image_count", "has_specs", "selected_for_extraction", "catalog_notes",
]


def merge_extraction(base: dict, override: dict) -> dict:
    """Copy non-empty extraction fields from override into base."""
    merged = dict(base)
    for field in _EXTRACTION_FIELDS:
        val = override.get(field)
        if val is not None and str(val).strip() not in ("", "None"):
            merged[field] = val
    # Track which sheet won for provenance
    merged["legacy_sheet_name"] = override.get("legacy_sheet_name", base.get("legacy_sheet_name"))
    merged["legacy_row_number"] = override.get("legacy_row_number", base.get("legacy_row_number"))
    return merged


def enrich_catalog(base: dict, catalog: dict) -> dict:
    """Overlay non-empty catalog fields from the Catalog Classification sheet."""
    enriched = dict(base)
    for field in _CATALOG_FIELDS:
        val = catalog.get(field)
        if val is not None and str(val).strip() not in ("", "None"):
            enriched[field] = val
    return enriched


# ---------------------------------------------------------------------------
# Migration orchestration
# ---------------------------------------------------------------------------

def run_migration(supabase_url: str, supabase_key: str, *, dry_run: bool) -> dict:
    sb = create_client(supabase_url, supabase_key)

    # ------------------------------------------------------------------
    # 1. Inventory all sheets
    # ------------------------------------------------------------------
    pipeline_sheets = list_sheets(PIPELINE_WORKBOOK)
    catalog_sheets  = list_sheets(CATALOG_WORKBOOK)

    has_master   = MASTER_SHEET   in pipeline_sheets
    has_enriched = ENRICHED_SHEET in pipeline_sheets
    has_failed   = FAILED_SHEET   in pipeline_sheets
    has_legacy   = LEGACY_SHEET   in pipeline_sheets
    has_catalog  = CATALOG_SHEET  in catalog_sheets

    # Provider tabs in the catalog workbook — skip, they are review artifacts
    provider_tabs = [
        s for s in catalog_sheets
        if s not in (CATALOG_SHEET, "Provider Tab Summary")
    ]

    print("=" * 60)
    print("PHASE 0A - SHEET INVENTORY")
    print("=" * 60)
    print(f"Pipeline workbook:  {PIPELINE_WORKBOOK.name}")
    for s in pipeline_sheets:
        flag = " [source]" if s in (MASTER_SHEET, ENRICHED_SHEET, FAILED_SHEET, LEGACY_SHEET) else " (skipped)"
        print(f"  {s}{flag}")
    print(f"\nCatalog workbook:   {CATALOG_WORKBOOK.name}")
    for s in catalog_sheets:
        flag = " [source]" if s == CATALOG_SHEET else f" (skipped - {len(provider_tabs)} provider tabs)" if s not in ("Provider Tab Summary",) else " (skipped)"
        print(f"  {s}{flag}")

    # ------------------------------------------------------------------
    # 2. Read all source sheets directly (no rebuild scripts)
    # ------------------------------------------------------------------
    print("\n" + "=" * 60)
    print("READING SOURCE SHEETS")
    print("=" * 60)

    master_rows   = read_sheet(PIPELINE_WORKBOOK, MASTER_SHEET)   if has_master   else []
    enriched_rows = read_sheet(PIPELINE_WORKBOOK, ENRICHED_SHEET) if has_enriched else []
    failed_rows   = read_sheet(PIPELINE_WORKBOOK, FAILED_SHEET)   if has_failed   else []
    legacy_rows   = read_sheet(PIPELINE_WORKBOOK, LEGACY_SHEET)   if has_legacy   else []
    catalog_rows  = read_sheet(CATALOG_WORKBOOK,  CATALOG_SHEET)  if has_catalog  else []

    print(f"  Contractors master:      {len(master_rows):>5} rows")
    print(f"  Contractors enriched:    {len(enriched_rows):>5} rows")
    print(f"  Request Failed:          {len(failed_rows):>5} rows")
    print(f"  Legacy Contractors:      {len(legacy_rows):>5} rows  {'(PRESENT)' if has_legacy else '(NOT FOUND)'}")
    print(f"  Catalog Classification:  {len(catalog_rows):>5} rows")

    # ------------------------------------------------------------------
    # 3. Build canonical index from Contractors master (baseline)
    # ------------------------------------------------------------------
    canonical:            dict[str, dict]  = {}
    master_collisions:    list[dict]       = []
    master_blank_company: list[int]        = []

    for row in master_rows:
        company    = str(row.get("Company") or "").strip()
        subcategory = str(row.get("Subcategory") or "").strip()

        if not company:
            master_blank_company.append(row["__row_number__"])
            continue

        mapped = map_row(row, source_sheet=MASTER_SHEET, workbook_name=PIPELINE_WORKBOOK.name)
        key = mapped["dedupe_key"]

        if key in canonical:
            master_collisions.append({
                "dedupe_key":          key,
                "company":             company,
                "subcategory":         subcategory,
                "collision_row":       row["__row_number__"],
                "first_seen_row":      canonical[key]["legacy_row_number"],
            })
            continue  # first occurrence wins
        canonical[key] = mapped

    # ------------------------------------------------------------------
    # 4. Merge Contractors enriched  (enriched > master for extraction fields)
    # ------------------------------------------------------------------
    enriched_merged:   int        = 0
    enriched_only:     int        = 0
    enriched_conflicts: list[dict] = []

    for row in enriched_rows:
        company     = str(row.get("Company") or "").strip()
        subcategory = str(row.get("Subcategory") or "").strip()
        if not company:
            continue

        mapped = map_row(row, source_sheet=ENRICHED_SHEET, workbook_name=PIPELINE_WORKBOOK.name)
        key = mapped["dedupe_key"]

        if key in canonical:
            existing_status = canonical[key].get("extraction_status", "")
            new_status      = mapped.get("extraction_status", "")
            # Flag but do not block: enriched always wins
            if existing_status == "request_failed" and new_status == "extraction_completed":
                enriched_conflicts.append({
                    "dedupe_key":       key,
                    "company":          company,
                    "subcategory":      subcategory,
                    "master_status":    existing_status,
                    "enriched_status":  new_status,
                })
            canonical[key] = merge_extraction(canonical[key], mapped)
            enriched_merged += 1
        else:
            # Row exists in enriched but not master — data integrity gap; import anyway
            canonical[key] = mapped
            enriched_only  += 1

    # ------------------------------------------------------------------
    # 5. Merge Request Failed (fills extraction fields where master has none)
    # ------------------------------------------------------------------
    failed_merged: int = 0
    failed_only:   int = 0

    for row in failed_rows:
        company     = str(row.get("Company") or "").strip()
        if not company:
            continue

        mapped = map_row(row, source_sheet=FAILED_SHEET, workbook_name=PIPELINE_WORKBOOK.name)
        key = mapped["dedupe_key"]

        if key in canonical:
            existing_status = canonical[key].get("extraction_status", "pending") or "pending"
            # Only update if master has no definitive extraction result yet
            if existing_status in ("pending", ""):
                canonical[key] = merge_extraction(canonical[key], mapped)
                failed_merged += 1
        else:
            # Row was triaged out of master entirely — import as-is
            canonical[key] = mapped
            failed_only    += 1

    # ------------------------------------------------------------------
    # 6. Enrich catalog fields from Catalog Classification
    # ------------------------------------------------------------------
    # Index catalog rows by (norm_company, norm_subcategory)
    catalog_index: dict[tuple[str, str], dict] = {}
    for row in catalog_rows:
        company     = str(row.get("Company") or "").strip()
        subcategory = str(row.get("Subcategory") or "").strip()
        if not company or not subcategory:
            continue
        cat_key = (norm_company(company), subcategory.lower())
        catalog_index[cat_key] = {
            "catalog_tier":            str(row.get("Catalog Tier") or "").strip() or None,
            "site_type":               str(row.get("Site Type") or "").strip() or None,
            "catalog_confirmed":       to_bool(row.get("Catalog Confirmed")),
            "product_subpages_found":  to_bool(row.get("Product Subpages Found")),
            "product_count_estimate":  to_int(row.get("Product Count Estimate")),
            "image_count":             to_int(row.get("Image Count")),
            "has_specs":               to_bool(row.get("Has Specs")),
            "selected_for_extraction": to_bool(row.get("Selected for Extraction")),
            "catalog_notes":           str(row.get("Notes") or "").strip() or None,
        }

    catalog_enriched:   int        = 0
    catalog_unmatched:  list[list] = []

    canonical_cat_keys = {
        (r["company"], (r.get("subcategory") or "").lower())
        for r in canonical.values()
    }

    for key, record in canonical.items():
        cat_key = (record["company"], (record.get("subcategory") or "").lower())
        if cat_key in catalog_index:
            canonical[key] = enrich_catalog(record, catalog_index[cat_key])
            catalog_enriched += 1

    catalog_unmatched = [
        list(k) for k in catalog_index if k not in canonical_cat_keys
    ]

    # ------------------------------------------------------------------
    # 7. Handle legacy Contractors sheet
    # ------------------------------------------------------------------
    legacy_added:   int        = 0
    legacy_skipped: int        = 0
    legacy_detail:  list[dict] = []

    for row in legacy_rows:
        # Old schema: Company may be in column "Company" or we fall back to "Contact Name"
        company     = str(row.get("Company") or "").strip()
        subcategory = str(row.get("Subcategory") or "").strip()
        if not company:
            company = str(row.get("Contact Name") or "").strip()
        if not company:
            continue

        mapped = map_row(row, source_sheet=LEGACY_SHEET, workbook_name=PIPELINE_WORKBOOK.name)
        key = mapped["dedupe_key"]

        if key in canonical:
            legacy_skipped += 1
            legacy_detail.append({"company": company, "subcategory": subcategory, "status": "skipped_duplicate"})
        else:
            canonical[key] = mapped
            legacy_added   += 1
            legacy_detail.append({"company": company, "subcategory": subcategory, "status": "added"})

    # ------------------------------------------------------------------
    # 8. Compute status breakdown and data quality flags
    # ------------------------------------------------------------------
    total_records = len(canonical)
    status_counts: dict[str, int] = {}
    blank_subcategory_keys: list[str] = []

    for key, record in canonical.items():
        status = record.get("extraction_status") or "pending"
        status_counts[status] = status_counts.get(status, 0) + 1
        if not record.get("subcategory"):
            blank_subcategory_keys.append(key)

    pending_count = status_counts.get("pending", 0)

    # ------------------------------------------------------------------
    # 9. Upsert to Supabase
    # ------------------------------------------------------------------
    print("\n" + "=" * 60)
    print(f"{'[DRY RUN] ' if dry_run else ''}UPSERTING {total_records} RECORDS TO SUPABASE")
    print("=" * 60)

    records  = list(canonical.values())
    upserted = 0
    errors:  list[dict] = []

    if dry_run:
        print(f"  [DRY RUN] Would upsert {total_records} records — no writes performed.")
        upserted = 0
    else:
        for i in range(0, len(records), UPSERT_BATCH):
            batch = records[i : i + UPSERT_BATCH]
            try:
                sb.table("contractor_rows").upsert(
                    batch, on_conflict="dedupe_key"
                ).execute()
                upserted += len(batch)
                print(f"  Batch {i // UPSERT_BATCH + 1:>3}: upserted {len(batch):>3} records  (total {upserted}/{total_records})")
            except Exception as exc:
                msg = str(exc)
                errors.append({"batch_start": i, "batch_size": len(batch), "error": msg})
                print(f"  ERROR batch {i // UPSERT_BATCH + 1}: {msg}")

    # ------------------------------------------------------------------
    # 10. Build and persist reconciliation report
    # ------------------------------------------------------------------
    report = {
        "generated_at":     datetime.now().isoformat(),
        "migration_batch_id": MIGRATION_BATCH_ID,
        "market_id":        MARKET_ID,
        "dry_run":          dry_run,

        "sheet_inventory": {
            "pipeline_workbook":              str(PIPELINE_WORKBOOK),
            "pipeline_sheets_found":          pipeline_sheets,
            "catalog_workbook":               str(CATALOG_WORKBOOK),
            "catalog_sheets_found":           catalog_sheets,
            "provider_tabs_found":            len(provider_tabs),
            "master_present":                 has_master,
            "enriched_present":               has_enriched,
            "failed_present":                 has_failed,
            "legacy_contractors_present":     has_legacy,
            "catalog_classification_present": has_catalog,
        },

        "source_row_counts": {
            "contractors_master":    len(master_rows),
            "contractors_enriched":  len(enriched_rows),
            "request_failed":        len(failed_rows),
            "legacy_contractors":    len(legacy_rows),
            "catalog_classification": len(catalog_rows),
        },

        "merge_results": {
            "unique_records_after_merge":       total_records,
            "master_internal_collisions":       len(master_collisions),
            "master_blank_company_rows":        master_blank_company,
            "enriched_merged_into_master":      enriched_merged,
            "enriched_only_records":            enriched_only,
            "failed_updated_existing":          failed_merged,
            "failed_only_records":              failed_only,
            "enriched_vs_failed_conflicts":     len(enriched_conflicts),
            "catalog_rows_enriched":            catalog_enriched,
            "catalog_unmatched_rows":           len(catalog_unmatched),
        },

        "legacy_sheet_treatment": {
            "sheet_present":          has_legacy,
            "rows_added":             legacy_added,
            "rows_skipped_duplicate": legacy_skipped,
        },

        "preservation": {
            "pending_extraction_rows":          pending_count,
            "extraction_completed_rows":        status_counts.get("extraction_completed", 0),
            "request_failed_rows":              status_counts.get("request_failed", 0),
            "no_contact_found_rows":            status_counts.get("no_contact_found", 0),
            "blank_extraction_status_rows":     status_counts.get("", 0),
        },

        "extraction_status_breakdown": status_counts,

        "data_quality": {
            "blank_subcategory_count":          len(blank_subcategory_keys),
            "blank_subcategory_dedupe_keys":    blank_subcategory_keys[:20],
            "master_blank_company_count":       len(master_blank_company),
            "master_blank_company_rows":        master_blank_company,
        },

        "ambiguous_dedup_cases": {
            "master_internal_collisions":       master_collisions[:50],
            "enriched_vs_failed_conflicts":     enriched_conflicts,
            "catalog_unmatched_keys":           catalog_unmatched[:50],
        },

        "import_result": {
            "upserted":  upserted,
            "errors":    errors,
        },

        "validation": {
            "source_master_rows":           len(master_rows),
            "supabase_records_upserted":    upserted,
            "delta":                        upserted - len(master_rows),
            "delta_note":                   "Delta >= 0 means enriched-only, failed-only, or legacy-only rows were added. Delta < 0 indicates an import error.",
            "status":                       "PASS" if (not errors and (dry_run or upserted > 0)) else "FAIL",
        },
    }

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = REPORT_DIR / f"migration_report_{ts}{'_dry' if dry_run else ''}.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    # ------------------------------------------------------------------
    # 11. Print human-readable summary
    # ------------------------------------------------------------------
    _print_summary(report)
    print(f"\nFull report: {report_path}")

    return report


def _print_summary(r: dict) -> None:
    src = r["source_row_counts"]
    mrg = r["merge_results"]
    leg = r["legacy_sheet_treatment"]
    prs = r["preservation"]
    imp = r["import_result"]
    val = r["validation"]
    dq  = r["data_quality"]

    print("\n" + "=" * 60)
    print("RECONCILIATION REPORT")
    print("=" * 60)

    print("\n--- SOURCE ROW COUNTS ---")
    print(f"  Contractors master:      {src['contractors_master']:>6}")
    print(f"  Contractors enriched:    {src['contractors_enriched']:>6}")
    print(f"  Request Failed:          {src['request_failed']:>6}")
    print(f"  Legacy Contractors:      {src['legacy_contractors']:>6}  {'(PRESENT)' if r['sheet_inventory']['legacy_contractors_present'] else '(NOT FOUND)'}")
    print(f"  Catalog Classification:  {src['catalog_classification']:>6}")

    print("\n--- MERGE RESULTS ---")
    print(f"  Unique canonical records:        {mrg['unique_records_after_merge']:>6}")
    print(f"  Master internal collisions:      {mrg['master_internal_collisions']:>6}  {'[!] review' if mrg['master_internal_collisions'] else ''}")
    print(f"  Enriched merged into master:     {mrg['enriched_merged_into_master']:>6}")
    print(f"  Enriched-only (not in master):   {mrg['enriched_only_records']:>6}  {'[!] integrity gap' if mrg['enriched_only_records'] else ''}")
    print(f"  Failed updated existing:         {mrg['failed_updated_existing']:>6}")
    print(f"  Failed-only (not in master):     {mrg['failed_only_records']:>6}")
    if mrg["enriched_vs_failed_conflicts"]:
        print(f"  [!] Enriched/failed conflicts:   {mrg['enriched_vs_failed_conflicts']:>6}  (enriched wins)")
    print(f"  Catalog fields enriched:         {mrg['catalog_rows_enriched']:>6}")
    print(f"  Catalog unmatched rows:          {mrg['catalog_unmatched_rows']:>6}")

    print("\n--- LEGACY CONTRACTORS SHEET ---")
    if r["sheet_inventory"]["legacy_contractors_present"]:
        print(f"  Rows added (net-new):            {leg['rows_added']:>6}")
        print(f"  Rows skipped (duplicate):        {leg['rows_skipped_duplicate']:>6}")
    else:
        print("  Sheet not present - skipped.")

    print("\n--- PRESERVATION (pending rows are kept intact) ---")
    print(f"  Pending extraction:              {prs['pending_extraction_rows']:>6}")
    print(f"  Extraction completed:            {prs['extraction_completed_rows']:>6}")
    print(f"  Request failed:                  {prs['request_failed_rows']:>6}")
    print(f"  No contact found:                {prs['no_contact_found_rows']:>6}")
    print(f"  Blank status:                    {prs['blank_extraction_status_rows']:>6}")

    if dq["blank_subcategory_count"] or dq["master_blank_company_count"]:
        print("\n--- DATA QUALITY WARNINGS ---")
        if dq["master_blank_company_count"]:
            print(f"  [!] Blank company names (skipped): {dq['master_blank_company_count']:>4}  rows {dq['master_blank_company_rows'][:5]}")
        if dq["blank_subcategory_count"]:
            print(f"  [!] Blank subcategory rows:        {dq['blank_subcategory_count']:>4}  (dedupe key uses company+market only)")

    print("\n--- SUPABASE IMPORT ---")
    if r["dry_run"]:
        print(f"  [DRY RUN] Records that would be upserted: {mrg['unique_records_after_merge']}")
    else:
        print(f"  Records upserted:                {imp['upserted']:>6}")
        print(f"  Import errors:                   {len(imp['errors']):>6}  {'[FAIL]' if imp['errors'] else '[OK]'}")

    print("\n--- VALIDATION ---")
    print(f"  Source master rows:              {val['source_master_rows']:>6}")
    print(f"  Supabase records upserted:       {val['supabase_records_upserted']:>6}")
    delta = val["delta"]
    sign  = "+" if delta > 0 else ""
    print(f"  Delta:                           {sign}{delta:>5}  {val['delta_note'][:45]}")
    status_icon = "[PASS]" if val["status"] == "PASS" else "[FAIL]"
    print(f"  Status:                                 {status_icon}")
    print("=" * 60)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    dry_run = "--dry-run" in sys.argv

    if dry_run:
        print("\n[DRY RUN MODE - Supabase will not be written to]\n")

    # Require source workbook
    if not PIPELINE_WORKBOOK.exists():
        print(f"ERROR: Pipeline workbook not found:\n  {PIPELINE_WORKBOOK}")
        return 1

    if not CATALOG_WORKBOOK.exists():
        print(f"WARNING: Catalog workbook not found — catalog enrichment will be skipped:\n  {CATALOG_WORKBOOK}")

    # Load credentials
    supabase_url = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "").strip()
    supabase_key = (
        os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
        or os.getenv("NEXT_PUBLIC_SUPABASE_ANON_KEY", "")
    ).strip()

    if not supabase_url or not supabase_key:
        print(
            "ERROR: Set NEXT_PUBLIC_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY "
            "in .env.local (project root)."
        )
        return 1

    # Prefer service role key — anon key will fail on INSERT
    if os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip():
        print("Using SUPABASE_SERVICE_ROLE_KEY.")
    else:
        print("WARNING: Using anon key — upserts may fail due to RLS. Set SUPABASE_SERVICE_ROLE_KEY.")

    try:
        report = run_migration(supabase_url, supabase_key, dry_run=dry_run)
    except Exception as exc:
        import traceback
        print(f"\nFATAL: {exc}")
        traceback.print_exc()
        return 1

    return 0 if report["validation"]["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
